# -*- coding: utf-8 -*-
# Author : Peter Wu

from odoo import models,fields,api
from odoo.exceptions import UserError
from datetime import datetime
import tempfile,base64,os
import openpyxl,shutil
from openpyxl.styles import Side, Border, colors
from odoo.http import request
#from base64 import b64decode,b64encode
from io import open


class CloudRentWeekReport(models.TransientModel):
    _name = "cloudrent.week_report"
    _description = "週報列印精靈(每週媒合成功案件統計表)"

    escrow_no = fields.Many2one('cloudrent.escrow',string="業者",default=lambda self:self.env.user.escrow_no.id)
    report_date = fields.Date(string="日期",default=datetime.today(),required=True)
    report_year = fields.Char(string="西元年:XXXX",required=True)
    week_seq = fields.Integer(string="週次",required=True)
    report_type = fields.Many2one('cloudrent.doc_filename',string="月報種類",required=True)

    @api.onchange('report_date')
    def onchangerdate(self):
        if not self.report_date:
            self.week_seq=''
            self.report_year=''
        else:
            self.env.cr.execute("""select getreportweek('%s')""" % self.report_date)
            self.week_seq = self.env.cr.fetchone()[0]
            self.env.cr.execute("""select getreportyear('%s')""" % self.report_date)
            self.report_year = self.env.cr.fetchone()[0]



    def run_week_report(self):
        from base64 import b64decode, b64encode
        border = Border(left=Side(border_style='thin', color ='000000'),right = Side(border_style='thin', color ='000000'),top = Side(border_style='thin', color ='000000'),bottom = Side(border_style='thin', color ='000000'))
        # 原 EXCEL 表格
        xlsx_path = self.env['ir.config_parameter'].sudo().get_param('cloudrent.xlsx.path')
        # copy 到目的端檔案
        xlsx_path1 = self.env['ir.config_parameter'].sudo().get_param('cloudrent.xlsx.path1')
        if xlsx_path[-1:] != "/":
            fn = xlsx_path + "/contract_match.xlsx"
        else:
            fn = xlsx_path + "contract_match.xlsx"
        self.env.cr.execute("""select get_xlsxtemp()""")
        tempfilename = self.env.cr.fetchone()[0]
        mytag = self.report_type.name1.contract_tag
        mydocrec = self.env['cloudrent.doc_filename'].search([('id', '=', self.report_type.id)])
        self.env.cr.execute("""select get_xlsxtemp()""")
        tempfilename = self.env.cr.fetchone()[0]
        if xlsx_path1[-1:] != "/":
            xlsxnewname = xlsx_path1 + "/" + tempfilename
        else:
            xlsxnewname = xlsx_path1 + tempfilename
        if mydocrec.doc_binfile:
            bfile = base64.b64decode(mydocrec.doc_binfile)
        elif mydocrec.doc_binfile1:
            bfile = base64.b64decode(mydocrec.doc_binfile1)
        with open(xlsxnewname, "wb") as f:
            f.write(bfile)
        if mytag == 17 :
            wb=openpyxl.load_workbook(xlsxnewname)
            myworksheet = 0
            #sheet=wb.active
            sheet = wb.worksheets[myworksheet]
            #ws=wb.active
            ws = wb.worksheets[myworksheet]
            ws.title = "第1頁"
            self.env.cr.execute("""select get_report_twdate('%s')""" % self.report_date)
            reportdate = self.env.cr.fetchone()[0]
            myrec = self.env['cloudrent.contract_match'].search([('match_year','=',self.report_year),('week_seq','=',self.week_seq),('escrow_no','=',self.escrow_no.id)])
            mycount = self.env['cloudrent.contract_match'].search_count([('match_year', '=', self.report_year), ('week_seq', '=', self.week_seq),('escrow_no','=',self.escrow_no.id)])
            mypage = 1
            mypagemod = divmod(mycount,15)
            if mypagemod[1] > 0 :
                mypage = mypagemod[0] + 1
            elif mypagemod[0] > 0 and mypagemod[1] == 0 :
                mypage = mypagemod[0]

            if mypage > 1:
                myitem = 16
                myseq = 5
                for i in range(mypage -1):
                    target = wb.copy_worksheet(ws)
                    target.title = "第" + str(i+2) + "頁"
                    for j in range(15):
                        target.cell(column=1, row=myseq).value = myitem
                        myseq = myseq + 1
                        myitem = myitem + 1

            ws.cell(column=1,row=2).value="公會編號: %s" % self.env.user.escrow_no.guild_no

            ws.merge_cells('Y2:AC2')
            top_left_cell = ws['Y2']
            top_left_cell.value = reportdate
            top_left_cell.border = border

            ws.merge_cells('A3:A4')
            top_left_cell1 = ws['A4']
            top_left_cell1.border = border
            ws.merge_cells('B3:B4')
            top_left_cell2 = ws['B4']
            top_left_cell2.border = border
            #ws.cell(column=26,row=2).value = reportdate
            nseq = 5
            mycount = 0
            for rec in myrec:
                A = divmod(mycount,15)                # 每 sheet 有15筆 紀錄 (週報 15 筆)
                if A[0] > 0 and A[1] == 1 :           # 換頁
                    myworksheet = myworksheet + 1
                    nseq = 5
                ws = wb.worksheets[myworksheet]
                busid = rec.escrow_no.id
                casetype=0      # 包租
                casetype1=0     # 代管
                if rec.case_type=='1':   # 包租
                    casetype=1
                elif rec.case_type=='2': # 代管
                    casetype1=1
                adminarea=''
                if rec.admin_area == '1':
                    adminarea = '台北市'
                elif rec.admin_area == '2':
                    adminarea = '新北市'
                elif rec.admin_area == '3':
                    adminarea = '桃園市'
                elif rec.admin_area == '4':
                    adminarea = '台中市'
                elif rec.admin_area == '5':
                    adminarea = '台南市'
                elif rec.admin_area == '6':
                    adminarea = '高雄市'
                genbuild=''
                genbuild1=''
                if rec.general_build == '1':   # 透天
                    genbuild = 1
                if rec.general_build == '2':   # 別墅
                    genbuild1 = 1
                btype = ''
                btype1 = ''
                btype2 = ''
                if rec.build_type == '1':        # 公寓
                    btype = 1
                elif rec.build_type == '2':      # 華廈
                    btype1 = 1
                elif rec.build_type == '3':      # 電梯大樓
                    btype2 = 1
                bpattern=''
                bpattern1=''
                bpattern2=''
                bpattern3=''
                if rec.build_pattern=='1':
                    bpattern=1
                elif rec.build_pattern=='2':
                    bpattern1=1
                elif rec.build_pattern=='3':
                    bpattern2=1
                elif rec.build_pattern=='4':
                    bpattern3=1

                lesseetype=''
                lesseetype1=''
                lesseetype2=''
                lesseetype3=''

                ##########

                ws.cell(column=2,row=nseq).value=rec.escrow_no.bus_name   # 業者名稱
                ws.cell(column=2,row=nseq).border = border
                ws.cell(column=3,row=nseq).value=rec.match_no             # 媒合編號
                ws.cell(column=3, row=nseq).border = border
                if casetype == 1 :    # 包租
                    ws.cell(column=4,row=nseq).value = 1
                else:
                    ws.cell(column=4, row=nseq).value = ''
                ws.cell(column=4, row=nseq).border = border
                if casetype1 == 1 :   # 代管
                    ws.cell(column=5,row=nseq).value = 1
                else:
                    ws.cell(column=5,row=nseq).value = ''
                ws.cell(column=5, row=nseq).border = border
                ws.cell(column=6,row=nseq).value = adminarea             # 物件行政區
                ws.cell(column=6, row=nseq).border = border
                ws.cell(column=7, row=nseq).value = rec.build_sec        # 段
                ws.cell(column=7, row=nseq).border = border
                ws.cell(column=8, row=nseq).value = rec.build_msec       # 小段
                ws.cell(column=8, row=nseq).border = border
                ws.cell(column=9, row=nseq).value = rec.build_number     # 建號
                ws.cell(column=9, row=nseq).border = border
                if rec.build_createdate:
                    ws.cell(column=10, row=nseq).value = rec.build_createdate.strftime("%Y-%m-%d")
                    ws.cell(column=10, row=nseq).border = border
                ws.cell(column=11, row=nseq).value = rec.build_age       # 屋齡
                ws.cell(column=11, row=nseq).border = border
                ws.cell(column=12, row=nseq).value = rec.build_area      # 面積
                ws.cell(column=12, row=nseq).border = border
                ws.cell(column=13, row=nseq).value = rec.build_for_rent  # 待租租金
                ws.cell(column=13, row=nseq).border = border
                ws.cell(column=14, row=nseq).value = rec.build_contract_rent # 簽約租金
                ws.cell(column=14, row=nseq).border = border
                if genbuild == 1:
                    ws.cell(column=15, row=nseq).value = 1               # 透天厝
                else:
                    ws.cell(column=15, row=nseq).value = ''
                ws.cell(column=15, row=nseq).border = border
                if genbuild1 == 1:
                    ws.cell(column=16, row=nseq).value = 1               # 別墅
                else:
                    ws.cell(column=16, row=nseq).value = ''
                ws.cell(column=16, row=nseq).border = border
                if btype == 1 :
                    ws.cell(column=17, row=nseq).value = 1               # 公寓
                else:
                    ws.cell(column=17, row=nseq).value = ''
                ws.cell(column=17, row=nseq).border = border
                if btype1 == 1 :
                    ws.cell(column=18, row=nseq).value = 1               # 華廈
                else:
                    ws.cell(column=18, row=nseq).value = ''
                ws.cell(column=18, row=nseq).border = border
                if btype2 == 1:
                    ws.cell(column=19, row=nseq).value = 1               # 電梯大樓
                else:
                    ws.cell(column=19, row=nseq).value = ''
                ws.cell(column=19, row=nseq).border = border
                if bpattern == 1 :
                    ws.cell(column=20, row=nseq).value = 1
                else:
                    ws.cell(column=20, row=nseq).value = ''
                ws.cell(column=20, row=nseq).border = border
                if bpattern1 == 1:
                    ws.cell(column=21, row=nseq).value = 1
                else:
                    ws.cell(column=21, row=nseq).value = ''
                ws.cell(column=21, row=nseq).border = border
                if bpattern2 == 1:
                    ws.cell(column=22, row=nseq).value = 1
                else:
                    ws.cell(column=22, row=nseq).value = ''
                ws.cell(column=22, row=nseq).border = border
                if bpattern3 == 1:
                    ws.cell(column=23, row=nseq).value = 1
                else:
                    ws.cell(column=23, row=nseq).value = ''
                ws.cell(column=23, row=nseq).border = border
                if rec.lessee_type == '1':
                    ws.cell(column=24, row=nseq).value = 1
                    ws.cell(column=24, row=nseq).border = border
                elif rec.lessee_type == '2':
                    ws.cell(column=26, row=nseq).value = 1
                    ws.cell(column=26, row=nseq).border = border
                elif rec.lessee_type == '3':
                    ws.cell(column=28, row=nseq).value = 1
                    ws.cell(column=28, row=nseq).border = border
                elif rec.lessee_type == '4':
                    ws.cell(column=25, row=nseq).value = 1
                    ws.cell(column=25, row=nseq).border = border
                if rec.is_send==True:
                    ws.cell(column=29, row=nseq).value = 'Y'
                    ws.cell(column=29, row=nseq).border = border

                nseq = nseq + 1
                mycount = mycount + 1



            wb.save(xlsxnewname)
            myxlsfilename = "[%s](%d週)媒合週報.xlsx" % (self.env.user.escrow_no.guild_no,self.week_seq)
            myrec = self.env['cloudrent.week_download']
            xlsxfile = open(xlsxnewname, "rb")
            myout = xlsxfile.read()
            myrec.create({'match_year':self.report_year,'week_seq':self.week_seq,'escrow_no':self.env.user.escrow_no.id,'xls_file': base64.b64encode(myout), 'xls_file_name': myxlsfilename})
            os.remove(xlsxnewname)
            myviewid = self.env.ref('cloudrent_escrow.view_cloudrent_week_download_tree')

            return {
                'view_name': 'CloudRentMatchWeekDownload',
                'name': ('媒合週報匯出'),
                'type': 'ir.actions.act_window',
                'res_model': 'cloudrent.week_download',
                'view_id': myviewid.id,
                'flags': {'action_buttons': True},
                'view_type': 'form',
                'view_mode': 'tree',
                'target': 'main'}



