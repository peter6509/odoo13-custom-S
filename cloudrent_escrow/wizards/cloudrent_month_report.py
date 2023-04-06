# -*- coding: utf-8 -*-
# Author : Peter Wu

from odoo import models,fields,api
from odoo.exceptions import UserError
from datetime import datetime
import tempfile,base64,os
import openpyxl,shutil
from openpyxl.styles import Side, Border, colors
from odoo.http import request
#from base64 import b64decode,b64encode
from io import open

class CloudRentMonthReport(models.TransientModel):
    _name = "cloudrent.month_report"
    _description = "月報列印精靈"

    escrow_no = fields.Many2one('cloudrent.escrow',string="業者",default=lambda self:self.env.user.escrow_no.id)
    report_date = fields.Date(string="製表日期", default=datetime.today(), required=True)
    report_year = fields.Char(string="西元年:XXXX",required=True)
    report_month = fields.Char(string="月份:MM",required=True)
    # report_type = fields.Selection([('1','出租人補助費用清單'),('2','承租人補助費用清單'),('3','業者補助費用申請書'),('4','業者補助費用清冊'),('5','業者關懷訪視月報'),('6','房屋租賃契約終止清冊')],string="月報種類")
    report_type = fields.Many2one('cloudrent.doc_filename',string="月報種類",required=True)

    def run_month_report(self):
        from base64 import b64decode,b64encode
        # 原 EXCEL 表格
        xlsx_path = self.env['ir.config_parameter'].sudo().get_param('cloudrent.xlsx.path')
        # copy 到目的端檔案
        xlsx_path1 = self.env['ir.config_parameter'].sudo().get_param('cloudrent.xlsx.path1')
        border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))

        # 定义边框样式
        def my_border(t_border, b_border, l_border, r_border):
            border = Border(top=Side(border_style=t_border, color=colors.BLACK),
                            bottom=Side(border_style=b_border, color=colors.BLACK),
                            left=Side(border_style=l_border, color=colors.BLACK),
                            right=Side(border_style=r_border, color=colors.BLACK))
            return border

        # 初始化制定区域边框为所有框线
        def format_border(s_column, s_index, e_column, e_index):
            for row in tuple(sheet[s_column + str(s_index):e_column + str(e_index)]):
                for cell in row:
                    cell.border = my_border('thin', 'thin', 'thin', 'thin')

        def set_solid_border(area_list):
            for area in area_list:
                s_column = area[0]
                s_index = area[1]
                e_column = area[2]
                e_index = area[3]
                # 设置左粗框线
                for cell in sheet[s_column][s_index - 1:e_index]:
                    cell.border = my_border(cell.border.top.style, cell.border.bottom.style,
                                            'thin', cell.border.right.style)
                # 设置右粗框线
                for cell in sheet[e_column][s_index - 1:e_index]:
                    cell.border = my_border(cell.border.top.style, cell.border.bottom.style,
                                            cell.border.left.style, 'thin')
                # 设置上粗框线
                for row in tuple(sheet[s_column + str(s_index):e_column + str(s_index)]):
                    for cell in row:
                        cell.border = my_border('thin', cell.border.bottom.style,
                                                cell.border.left.style, cell.border.right.style)
                # 设置下粗框线
                for row in tuple(sheet[s_column + str(e_index):e_column + str(e_index)]):
                    for cell in row:
                        cell.border = my_border(cell.border.top.style, 'thin',
                                                cell.border.left.style, cell.border.right.style)
        mytag = self.report_type.name1.contract_tag

        mydocrec = self.env['cloudrent.doc_filename'].search([('id', '=', self.report_type.id)])
        self.env.cr.execute("""select get_xlsxtemp()""")
        tempfilename = self.env.cr.fetchone()[0]
        if xlsx_path1[-1:] != "/":
            xlsxnewname = xlsx_path1 + "/" + tempfilename
        else:
            xlsxnewname = xlsx_path1 + tempfilename
        if mydocrec.doc_binfile:
            bfile = base64.b64decode(mydocrec.doc_binfile)
        elif mydocrec.doc_binfile1:
            bfile = base64.b64decode(mydocrec.doc_binfile1)
        with open(xlsxnewname, "wb") as f:
            f.write(bfile)
        if mytag == 11 :    # 出租人補助費用清單  10 筆
            # if xlsx_path[-1:] != "/":
            #     fn = xlsx_path + "/lessor_grant_fee.xlsx"
            # else:
            #     fn = xlsx_path + "lessor_grant_fee.xlsx"
            # self.env.cr.execute("""select get_xlsxtemp()""")
            # tempfilename = self.env.cr.fetchone()[0]

            #
            # shutil.copyfile(fn, xlsxnewname)

            wb = openpyxl.load_workbook(xlsxnewname)
            sheet = wb.active
            ws = wb.active
            ws.title = "第1頁"
            self.env.cr.execute("""select get_report_twdate('%s')""" % self.report_date)
            reportdate = self.env.cr.fetchone()[0]
            myrec = self.env['cloudrent.contract_match'].search([('match_year', '=', self.report_year), ('match_month', '=', self.report_month),('escrow_no','=',self.escrow_no.id)])
            mycount = self.env['cloudrent.contract_match'].search_count([('match_year', '=', self.report_year), ('match_month', '=', self.report_month),('escrow_no','=',self.escrow_no.id)])
            #mycount = self.env['cloudrent.contract_match'].search_count([('match_year', '=', self.report_year), ('week_seq', '=', self.week_seq)])

            mypage = 1
            mypagemod = divmod(mycount, 10)
            if mypagemod[1] > 0:
                mypage = mypagemod[0] + 1
            elif mypagemod[0] > 0 and mypagemod[1] == 0:
                mypage = mypagemod[0]

            if mypage > 1:
                myitem = 11
                for i in range(mypage - 1):
                    target = wb.copy_worksheet(ws)
                    target.title = "第" + str(i + 2) + "頁"
                    myseq = 5
                    for j in range(10):
                        target.cell(column=1, row=myseq).value = myitem
                        myseq = myseq + 1
                        myitem = myitem + 1



            ws.merge_cells('A2:B2')
            top_left_cell = ws['A2']
            ws.cell(column=1, row=2).value = "業者編號: %s" % self.env.user.escrow_no.bus_no
            ws.merge_cells('L2:N2')
            top_left_cell1 = ws['L2']
            ws.cell(column=12, row=2).value = reportdate
            nseq = 5
            mycount = 0
            myworksheet = 0
            for rec in myrec:
                A = divmod(mycount, 10)  # 每 sheet 有10筆 紀錄 (週報 10 筆)
                if A[0] > 0 and A[1] == 1:  # 換頁
                    myworksheet = myworksheet + 1
                    nseq = 5
                ws = wb.worksheets[myworksheet]
                ws.cell(column=2, row=nseq).value = rec.match_no  # 媒合編號
                ws.cell(column=2, row=nseq).border = border

                self.env.cr.execute("""select get_lessorgrant(%d,'%s','%s','%s')""" % (rec.id,self.report_year,self.report_month,'S'))
                mydata = self.env.cr.fetchall()
                if mydata and mydata[0] > 0:
                    ws.cell(column=3, row=nseq).value = mydata[0]     # 保險費 實際金額
                ws.cell(column=3, row=nseq).border = border
                if mydata and mydata[1] > 0:
                    ws.cell(column=4, row=nseq).value = mydata[1]  # 保險費 申請金額
                ws.cell(column=4, row=nseq).border = border

                self.env.cr.execute("""select get_lessorgrant(%d,'%s','%s','%s')""" % (rec.id, self.report_year, self.report_month, 'G'))
                mydata = self.env.cr.fetchall()
                if mydata and mydata[0] > 0:
                    ws.cell(column=5, row=nseq).value = mydata[0]  # 公證費 實際金額
                ws.cell(column=5, row=nseq).border = border
                if mydata and mydata[1] > 0:
                    ws.cell(column=6, row=nseq).value = mydata[1]  # 公證費 申請金額
                ws.cell(column=6, row=nseq).border = border

                self.env.cr.execute("""select get_lessorgrant(%d,'%s','%s','%s')""" % (rec.id, self.report_year, self.report_month, 'R'))
                mydata = self.env.cr.fetchall()
                if mydata and mydata[0] > 0:
                    ws.cell(column=7, row=nseq).value = mydata[0]  # 修繕費 實際金額
                ws.cell(column=7, row=nseq).border = border
                if mydata and mydata[1] > 0:
                    ws.cell(column=8, row=nseq).value = mydata[1]  # 修繕費 申請金額
                ws.cell(column=8, row=nseq).border = border

                ws.cell(column=9, row=nseq).value = rec.build_lessor.escrow_man    # 受款人 姓名
                ws.cell(column=9, row=nseq).border = border

                ws.cell(column=10, row=nseq).value = rec.lessor_pid  # 受款人 身份證號
                ws.cell(column=10, row=nseq).border = border

                ws.cell(column=11, row=nseq).value = rec.lessor_fin_instno  # 受款人 金融機構代碼
                ws.cell(column=11, row=nseq).border = border

                ws.cell(column=12, row=nseq).value = rec.lessor_fin_branch  # 受款人 分行
                ws.cell(column=12, row=nseq).border = border

                ws.cell(column=13, row=nseq).value = rec.lessor_fin_account  # 受款人 帳號
                ws.cell(column=13, row=nseq).border = border

                nseq = nseq + 1
                mycount = mycount + 1

            wb.save(xlsxnewname)
            myxlsfilename = "[%s]%s-%s出租人補助費用清單.xlsx" % (
            self.env.user.escrow_no.bus_no, self.report_year, self.report_month)
            myrec1 = self.env['cloudrent.month_download']
            xlsxfile = open(xlsxnewname, "rb")
            myout = xlsxfile.read()
            myrec1.create({'match_year': self.report_year, 'escrow_no': self.env.user.escrow_no.id,
                          'xls_file': base64.b64encode(myout),'match_month':self.report_month , 'xls_file_name': myxlsfilename})
            os.remove(xlsxnewname)
            myviewid = self.env.ref('cloudrent_escrow.view_cloudrent_month_download_tree')

            return {
                'view_name': 'CloudRentMatchMonthDownload',
                'name': ('出租人補助費用清單匯出'),
                'type': 'ir.actions.act_window',
                'res_model': 'cloudrent.month_download',
                'view_id': myviewid.id,
                'flags': {'action_buttons': True},
                'view_type': 'form',
                'view_mode': 'tree',
                'target': 'main'}
        # elif self.report_type == '2':
        elif mytag == 12 :                 # 承租人補助費用清單 10筆
            # if xlsx_path[-1:] != "/":
            #     fn = xlsx_path + "/lessee_grant_fee.xlsx"
            # else:
            #     fn = xlsx_path + "lessee_grant_fee.xlsx"
            # self.env.cr.execute("""select get_xlsxtemp()""")
            # tempfilename = self.env.cr.fetchone()[0]
            # if xlsx_path1[-1:] != "/":
            #     xlsxnewname = xlsx_path1 + "/" + tempfilename
            # else:
            #     xlsxnewname = xlsx_path1 + tempfilename
            #
            # shutil.copyfile(fn, xlsxnewname)
            wb = openpyxl.load_workbook(xlsxnewname)
            sheet = wb.active
            ws = wb.active
            ws.title = "第1頁"
            self.env.cr.execute("""select get_report_twdate('%s')""" % self.report_date)
            reportdate = self.env.cr.fetchone()[0]
            myrec = self.env['cloudrent.contract_match'].search([('match_year', '=', self.report_year), ('match_month', '=', self.report_month),('escrow_no','=',self.escrow_no.id)])
            mycount = self.env['cloudrent.contract_match'].search_count([('match_year', '=', self.report_year), ('match_month', '=', self.report_month),('escrow_no','=',self.escrow_no.id)])

            mypage=1
            mypagemod = divmod(mycount, 10)
            if mypagemod[1] > 0:
                mypage = mypagemod[0] + 1
            elif mypagemod[0] > 0 and mypagemod[1] == 0:
                mypage = mypagemod[0]

            if mypage > 1:
                myitem = 11
                for i in range(mypage - 1):
                    target = wb.copy_worksheet(ws)
                    target.title = "第" + str(i + 2) + "頁"
                    myseq = 5
                    for j in range(10):
                        target.cell(column=1, row=myseq).value = myitem
                        myseq = myseq + 1
                        myitem = myitem + 1

            ws.merge_cells('A2:B2')
            top_left_cell = ws['A2']
            ws.cell(column=1, row=2).value = "業者編號: %s" % self.env.user.escrow_no.bus_no
            ws.merge_cells('L2:P2')
            top_left_cell1 = ws['L2']
            ws.cell(column=12, row=2).value = reportdate
            nseq = 5
            mycount = 0
            myworksheet = 0
            for rec in myrec:
                A = divmod(mycount, 10)  # 每 sheet 有10筆 紀錄 (週報 10 筆)
                if A[0] > 0 and A[1] == 1:  # 換頁
                    myworksheet = myworksheet + 1
                    nseq = 5
                ws = wb.worksheets[myworksheet]
                ws.cell(column=2, row=nseq).value = rec.match_no  # 媒合編號
                ws.cell(column=2, row=nseq).border = border

                self.env.cr.execute("""select get_lessorgrant(%d,'%s','%s','%s')""" % (rec.id,self.report_year,self.report_month,'S'))
                mydata = self.env.cr.fetchall()
                if mydata and mydata[0] > 0:
                    ws.cell(column=3, row=nseq).value = mydata[0]     # 保險費 實際金額
                ws.cell(column=3, row=nseq).border = border
                if mydata and mydata[1] > 0:
                    ws.cell(column=4, row=nseq).value = mydata[1]  # 保險費 申請金額
                ws.cell(column=4, row=nseq).border = border

                self.env.cr.execute("""select get_lessorgrant(%d,'%s','%s','%s')""" % (rec.id, self.report_year, self.report_month, 'G'))
                mydata = self.env.cr.fetchall()
                if mydata and mydata[0] > 0:
                    ws.cell(column=5, row=nseq).value = mydata[0]  # 公證費 實際金額
                ws.cell(column=5, row=nseq).border = border
                if mydata and mydata[1] > 0:
                    ws.cell(column=6, row=nseq).value = mydata[1]  # 公證費 申請金額
                ws.cell(column=6, row=nseq).border = border

                self.env.cr.execute("""select get_lessorgrant(%d,'%s','%s','%s')""" % (rec.id, self.report_year, self.report_month, 'R'))
                mydata = self.env.cr.fetchall()
                if mydata and mydata[0] > 0:
                    ws.cell(column=7, row=nseq).value = mydata[0]  # 修繕費 實際金額
                ws.cell(column=7, row=nseq).border = border
                if mydata and mydata[1] > 0:
                    ws.cell(column=8, row=nseq).value = mydata[1]  # 修繕費 申請金額
                ws.cell(column=8, row=nseq).border = border

                ws.cell(column=9, row=nseq).value = rec.build_lessor.escrow_man    # 受款人 姓名
                ws.cell(column=9, row=nseq).border = border

                ws.cell(column=10, row=nseq).value = rec.lessor_pid  # 受款人 身份證號
                ws.cell(column=10, row=nseq).border = border

                ws.cell(column=11, row=nseq).value = rec.lessor_fin_instno  # 受款人 金融機構代碼
                ws.cell(column=11, row=nseq).border = border

                ws.cell(column=12, row=nseq).value = rec.lessor_fin_branch  # 受款人 分行
                ws.cell(column=12, row=nseq).border = border

                ws.cell(column=13, row=nseq).value = rec.lessor_fin_account  # 受款人 帳號
                ws.cell(column=13, row=nseq).border = border

                nseq = nseq + 1
                mycount = mycount + 1

            wb.save(xlsxnewname)
            myxlsfilename = "[%s]%s-%s出租人補助費用清單.xlsx" % (
            self.env.user.escrow_no.bus_no, self.report_year, self.report_month)
            myrec1 = self.env['cloudrent.month_download']
            xlsxfile = open(xlsxnewname, "rb")
            myout = xlsxfile.read()
            myrec1.create({'match_year': self.report_year, 'escrow_no': self.env.user.escrow_no.id,
                          'xls_file': base64.b64encode(myout),'match_month':self.report_month , 'xls_file_name': myxlsfilename})
            os.remove(xlsxnewname)
            myviewid = self.env.ref('cloudrent_escrow.view_cloudrent_month_download_tree')

            return {
                'view_name': 'CloudRentMatchMonthDownload',
                'name': ('出租人補助費用清單匯出'),
                'type': 'ir.actions.act_window',
                'res_model': 'cloudrent.month_download',
                'view_id': myviewid.id,
                'flags': {'action_buttons': True},
                'view_type': 'form',
                'view_mode': 'tree',
                'target': 'main'}
        # elif self.report_type == '3':   # 業者服務補助費用申請書
        elif mytag == 13 :       # 業者服務補助費用申請書
            self.env.cr.execute("""select gen_agent_applyfor_report(%d,'%s','%s','%s')""" % (self.env.user.escrow_no.id,self.report_date,self.report_year,self.report_month))
            self.env.cr.execute("""commit""")
            myrec = self.env['cloudrent.agent_applyfor_report'].search([('escrow_no','=',self.env.user.escrow_no.id)])
            bf_url = request.env['ir.config_parameter'].sudo().get_param('web.bf.url')
            url = "%s/report/odt_to_x/report_agent_applyfor_report/%s" % (bf_url, myrec.id)
            return {'name': 'Go to website',
                    'res_model': 'ir.actions.act_url',
                    'type': 'ir.actions.act_url',
                    'target': 'new',
                    'url': url
                    }
        elif mytag == 14 :   # 業者補助費用清冊 10筆
            # if xlsx_path[-1:] != "/":
            #     fn = xlsx_path + "/agent_grant_fee.xlsx"
            # else:
            #     fn = xlsx_path + "agent_grant_fee.xlsx"
            # self.env.cr.execute("""select get_xlsxtemp()""")
            # tempfilename = self.env.cr.fetchone()[0]
            # if xlsx_path1[-1:] != "/":
            #     xlsxnewname = xlsx_path1 + "/" + tempfilename
            # else:
            #     xlsxnewname = xlsx_path1 + tempfilename
            #
            # shutil.copyfile(fn, xlsxnewname)
            wb = openpyxl.load_workbook(xlsxnewname)
            sheet = wb.active
            ws = wb.active
            ws.title = "第1頁"
            self.env.cr.execute("""select get_report_twdate('%s')""" % self.report_date)
            reportdate = self.env.cr.fetchone()[0]
            myrec = self.env['cloudrent.contract_match'].search([('match_year', '=', self.report_year), ('match_month', '=', self.report_month),('escrow_no','=',self.escrow_no.id)])
            mycount = self.env['cloudrent.contract_match'].search_count([('match_year', '=', self.report_year), ('match_month', '=', self.report_month),('escrow_no','=',self.escrow_no.id)])

            mypage = 1
            mypagemod = divmod(mycount, 10)
            if mypagemod[1] > 0:
                mypage = mypagemod[0] + 1
            elif mypagemod[0] > 0 and mypagemod[1] == 0:
                mypage = mypagemod[0]

            if mypage > 1:
                myitem = 11
                for i in range(mypage - 1):
                    target = wb.copy_worksheet(ws)
                    target.title = "第" + str(i + 2) + "頁"
                    myseq = 5
                    for j in range(10):
                        target.cell(column=1, row=myseq).value = myitem
                        myseq = myseq + 1
                        myitem = myitem + 1

            ws.merge_cells('A2:B2')
            top_left_cell = ws['A2']
            ws.cell(column=1, row=2).value = "公會編號: %s" % self.env.user.escrow_no.guild_no
            ws.merge_cells('L2:Q2')
            top_left_cell1 = ws['L2']
            ws.cell(column=12, row=2).value = reportdate
            nseq = 5
            mycount = 0
            myworksheet=0
            for rec in myrec:
                A = divmod(mycount, 10)  # 每 sheet 有10筆 紀錄 (週報 10 筆)
                if A[0] > 0 and A[1] == 1:  # 換頁
                    myworksheet = myworksheet + 1
                    nseq = 5
                ws = wb.worksheets[myworksheet]
                ws.cell(column=2, row=nseq).value = rec.match_no  # 媒合編號
                ws.cell(column=2, row=nseq).border = border
                if rec.lessee_type == '1':
                    mylesseetype = '一般戶'
                elif rec.lessee_type == '2':
                    mylesseetype = '一類戶'
                elif rec.lessee_type == '3':
                    mylesseetype = '二類戶'
                else:
                    mylesseetype = '就學,就業,警消'
                ws.cell(column=3, row=nseq).value = mylesseetype  # 承租人身份
                ws.cell(column=3, row=nseq).border = border

                ws.cell(column=4, row=nseq).value = rec.build_contract_rent  # 簽約租金
                ws.cell(column=4, row=nseq).border = border
                if rec.match_start_date:
                    self.env.cr.execute("""select get_twymd('%s')""" % rec.match_start_date)
                    ws.cell(column=5, row=nseq).value = self.env.cr.fetchone()[0]  # 簽約起日
                ws.cell(column=5, row=nseq).border = border
                if rec.match_end_date:
                    self.env.cr.execute("""select get_twymd('%s')""" % rec.match_end_date)
                    ws.cell(column=6, row=nseq).value = self.env.cr.fetchone()[0]  # 簽約迄日
                ws.cell(column=6, row=nseq).border = border

                ws.cell(column=7, row=nseq).value = ''  # 既存租約
                ws.cell(column=7, row=nseq).border = border

                self.env.cr.execute("""select get_ndmfee(%d,'%s','%s','%s')""" % (rec.id, self.report_year, self.report_month, 'N'))
                myres = self.env.cr.fetchone()[0]
                if myres > 0:
                    ws.cell(column=8, row=nseq).value = myres  # 公證費
                ws.cell(column=8, row=nseq).border = border

                self.env.cr.execute("""select get_ndmfee(%d,'%s','%s','%s')""" % (rec.id, self.report_year, self.report_month, 'D'))
                myres = self.env.cr.fetchone()[0]
                if myres > 0:
                    ws.cell(column=9, row=nseq).value = myres  # 開發費
                ws.cell(column=9, row=nseq).border = border

                self.env.cr.execute("""select get_gefee(%d,'%s','%s','%s')""" % (rec.id, self.report_year, self.report_month, 'G'))
                mydata = self.env.cr.fetchall()
                if mydata and mydata[0] > 0:
                    ws.cell(column=10, row=nseq).value = mydata[0]  # 包管費金額
                    ws.cell(column=11, row=nseq).value = mydata[1]  # 包管費期數
                    ws.cell(column=12, row=nseq).value = mydata[2]  # 包管費總期數
                ws.cell(column=10, row=nseq).border = border
                ws.cell(column=11, row=nseq).border = border
                ws.cell(column=12, row=nseq).border = border

                self.env.cr.execute("""select get_ndmfee(%d,'%s','%s','%s')""" % (rec.id, self.report_year, self.report_month, 'M'))
                myres = self.env.cr.fetchone()[0]
                if myres > 0:
                    ws.cell(column=13, row=nseq).value = myres  # 媒合費
                ws.cell(column=13, row=nseq).border = border

                self.env.cr.execute("""select get_gefee(%d,'%s','%s','%s')""" % (rec.id, self.report_year, self.report_month, 'E'))
                mydata = self.env.cr.fetchall()
                if mydata and mydata[0] > 0:
                    ws.cell(column=14, row=nseq).value = mydata[0]  # 代管費金額
                    ws.cell(column=15, row=nseq).value = mydata[1]  # 代管費期數
                    ws.cell(column=16, row=nseq).value = mydata[2]  # 代管費總期數
                ws.cell(column=14, row=nseq).border = border
                ws.cell(column=15, row=nseq).border = border
                ws.cell(column=16, row=nseq).border = border
                nseq = nseq + 1
                mycount = mycount + 1

            wb.save(xlsxnewname)
            myxlsfilename = "[%s]%s-%s業者補助費用清冊.xlsx" % (
            self.env.user.escrow_no.guild_no, self.report_year, self.report_month)
            myrec1 = self.env['cloudrent.month_download']
            xlsxfile = open(xlsxnewname, "rb")
            myout = xlsxfile.read()
            myrec1.create({'match_year': self.report_year, 'escrow_no': self.env.user.escrow_no.id,
                          'xls_file': base64.b64encode(myout),'match_month':self.report_month , 'xls_file_name': myxlsfilename})
            os.remove(xlsxnewname)
            myviewid = self.env.ref('cloudrent_escrow.view_cloudrent_month_download_tree')

            return {
                'view_name': 'CloudRentMatchMonthDownload',
                'name': ('業者補助費用清冊匯出'),
                'type': 'ir.actions.act_window',
                'res_model': 'cloudrent.month_download',
                'view_id': myviewid.id,
                'flags': {'action_buttons': True},
                'view_type': 'form',
                'view_mode': 'tree',
                'target': 'main'}

        elif mytag == 15 :  # 業者關懷訪視月報
            # if xlsx_path[-1:] != "/":
            #     fn = xlsx_path + "/agent_visit_service.xlsx"
            # else:
            #     fn = xlsx_path + "agent_visit_service.xlsx"
            # self.env.cr.execute("""select get_xlsxtemp()""")
            # tempfilename = self.env.cr.fetchone()[0]
            # if xlsx_path1[-1:] != "/":
            #     xlsxnewname = xlsx_path1 + "/" + tempfilename
            # else:
            #     xlsxnewname = xlsx_path1 + tempfilename
            #
            # shutil.copyfile(fn, xlsxnewname)
            wb = openpyxl.load_workbook(xlsxnewname)
            sheet = wb.active
            ws = wb.active
            ws.title = "第1頁"
            self.env.cr.execute("""select get_report_twdate('%s')""" % self.report_date)
            reportdate = self.env.cr.fetchone()[0]
            myrec = self.env['cloudrent.agent_visit_service'].search([('visit_year', '=', self.report_year), ('visit_month', '=', self.report_month),('escrow_no','=',self.escrow_no.id)])
            mycount = self.env['cloudrent.agent_visit_service'].search_count([('visit_year', '=', self.report_year), ('visit_month', '=', self.report_month),('escrow_no','=',self.escrow_no.id)])

            mypage = 1
            mypagemod = divmod(mycount, 4)
            if mypagemod[1] > 0:
                mypage = mypagemod[0] + 1
            elif mypagemod[0] > 0 and mypagemod[1] == 0:
                mypage = mypagemod[0]

            if mypage > 1:
                myitem = 5
                for i in range(mypage - 1):
                    target = wb.copy_worksheet(ws)
                    target.title = "第" + str(i + 2) + "頁"
                    myseq = 5
                    for j in range(4):
                        target.cell(column=1, row=myseq).value = myitem
                        myseq = myseq + 1
                        myitem = myitem + 1

            ws.merge_cells('A2:C2')
            top_left_cell = ws['A2']
            ws.cell(column=1, row=2).value = "業者編號: %s" % self.env.user.escrow_no.bus_no
            ws.merge_cells('K2:M2')
            top_left_cell1 = ws['K2']
            ws.cell(column=11, row=2).value = reportdate
            nseq = 5
            mycount = 0
            myworksheet=0
            for rec in myrec:
                A = divmod(mycount, 4)  # 每 sheet 有4筆 紀錄
                if A[0] > 0 and A[1] == 1:  # 換頁
                    myworksheet = myworksheet + 1
                    nseq = 5
                ws = wb.worksheets[myworksheet]
                ws.cell(column=3, row=nseq).value = rec.match_no.match_no  # 媒合編號
                ws.cell(column=3, row=nseq).border = border
                if rec.match_no.case_type=='1':   # 包租包管
                    mycasetype = '■包租包管\n□代租代管'
                elif rec.match_no.case_type=='2':
                    mycasetype = '□包租包管\n■代租代管'
                else:
                    mycasetype = '□包租包管\n□代租代管'
                ws.cell(column=4, row=nseq).value = mycasetype  # 案件類型
                ws.cell(column=4, row=nseq).border = border

                ws.cell(column=5, row=nseq).value = rec.match_no.build_lessee.escrow_man  # 承租人姓名
                ws.cell(column=5, row=nseq).border = border
                if rec.visit_date:
                    self.env.cr.execute("""select get_twymd('%s')""" % rec.visit_date)
                    ws.cell(column=6, row=nseq).value = self.env.cr.fetchone()[0]       # 訪視日期
                ws.cell(column=6, row=nseq).border = border
                if rec.visit_type=='1':
                    myvtype = '■電訪\n□家訪\n□其他'
                elif rec.visit_type=='2':
                    mytype = '□電訪\n■家訪\n□其他'
                elif rec.visit_type=='3':
                    mytype = '□電訪\n□家訪\n■其他'
                else:
                    mytype = '□電訪\n□家訪\n□其他'
                ws.cell(column=7, row=nseq).value = myvtype  # 訪視方式
                ws.cell(column=7, row=nseq).border = border

                if rec.require_item=='1':
                    myritem = '■無特殊情形\n□可能無力支付租金\n□有其他情形'
                elif rec.require_item=='2':
                    myritem = '□無特殊情形\n■可能無力支付租金\n□有其他情形'
                elif rec.require_item=='3':
                    myritem = '□無特殊情形\n□可能無力支付租金\n■有其他情形'
                else:
                    myritem = '□無特殊情形\n□可能無力支付租金\n□有其他情形'

                ws.cell(column=8, row=nseq).value = myritem  # 必要項目
                ws.cell(column=8, row=nseq).border = border

                if rec.other_alert=='1':
                    myalert = '■房客有急難或身心狀況不佳\n□違法行為疑慮\n□入住與申請資料不符之疑慮\n□不當使用或裝修\n□住宅安全疑慮\n□其他'
                elif rec.other_alert=='2':
                    myalert = '□房客有急難或身心狀況不佳\n■違法行為疑慮\n□入住與申請資料不符之疑慮\n□不當使用或裝修\n□住宅安全疑慮\n□其他'
                elif rec.other_alert=='3':
                    myalert = '□房客有急難或身心狀況不佳\n□違法行為疑慮\n■入住與申請資料不符之疑慮\n□不當使用或裝修\n□住宅安全疑慮\n□其他'
                elif rec.other_alert=='4':
                    myalert = '□房客有急難或身心狀況不佳\n□違法行為疑慮\n□入住與申請資料不符之疑慮\n■不當使用或裝修\n□住宅安全疑慮\n□其他'
                elif rec.other_alert=='5':
                    myalert = '□房客有急難或身心狀況不佳\n□違法行為疑慮\n□入住與申請資料不符之疑慮\n□不當使用或裝修\n■住宅安全疑慮\n□其他'
                elif rec.other_alert=='6':
                    myalert = '□房客有急難或身心狀況不佳\n□違法行為疑慮\n□入住與申請資料不符之疑慮\n□不當使用或裝修\n□住宅安全疑慮\n■其他'
                else:
                    myalert = '□房客有急難或身心狀況不佳\n□違法行為疑慮\n□入住與申請資料不符之疑慮\n□不當使用或裝修\n□住宅安全疑慮\n□其他'

                ws.cell(column=9, row=nseq).value = myalert  # 其他情形通報
                ws.cell(column=9, row=nseq).border = border

                ws.cell(column=10, row=nseq).value = rec.visit_memo  # 記錄摘要
                ws.cell(column=10, row=nseq).border = border
                ws.cell(column=11, row=nseq).value = rec.visit_process  # 處理情形
                ws.cell(column=11, row=nseq).border = border

                if rec.state=='1':
                    mystate = '■處理中\n□已結案'
                elif rec.state=='2':
                    mystate = '□處理中\n■已結案'
                else:
                    mystate = '□處理中\n□已結案'

                ws.cell(column=12, row=nseq).value = mystate     # 處理進度
                ws.cell(column=12, row=nseq).border = border

                nseq = nseq + 1
                mycount = mycount + 1

            wb.save(xlsxnewname)
            myxlsfilename = "[%s]%s-%s訪視記錄月報.xlsx" % (self.env.user.escrow_no.guild_no, self.report_year, self.report_month)
            myrec1 = self.env['cloudrent.month_download']
            xlsxfile = open(xlsxnewname, "rb")
            myout = xlsxfile.read()
            myrec1.create({'match_year': self.report_year, 'escrow_no': self.env.user.escrow_no.id,
                          'xls_file': base64.b64encode(myout),'match_month':self.report_month , 'xls_file_name': myxlsfilename})
            os.remove(xlsxnewname)
            myviewid = self.env.ref('cloudrent_escrow.view_cloudrent_month_download_tree')

            return {
                'view_name': 'CloudRentMatchMonthDownload',
                'name': ('業者訪視記錄月報匯出'),
                'type': 'ir.actions.act_window',
                'res_model': 'cloudrent.month_download',
                'view_id': myviewid.id,
                'flags': {'action_buttons': True},
                'view_type': 'form',
                'view_mode': 'tree',
                'target': 'main'}

        elif mytag == 16 :  # 房屋租賃契約終止清冊
            # if xlsx_path[-1:] != "/":
            #     fn = xlsx_path + "/contract_termination.xlsx"
            # else:
            #     fn = xlsx_path + "contract_termination.xlsx"
            # self.env.cr.execute("""select get_xlsxtemp()""")
            # tempfilename = self.env.cr.fetchone()[0]
            # if xlsx_path1[-1:] != "/":
            #     xlsxnewname = xlsx_path1 + "/" + tempfilename
            # else:
            #     xlsxnewname = xlsx_path1 + tempfilename
            #
            # shutil.copyfile(fn, xlsxnewname)
            wb = openpyxl.load_workbook(xlsxnewname)
            sheet = wb.active
            ws = wb.active
            ws.title = "第1頁"
            self.env.cr.execute("""select get_report_twdate('%s')""" % self.report_date)
            reportdate = self.env.cr.fetchone()[0]
            myrec = self.env['cloudrent.quit_lease'].search([('quit_year', '=', self.report_year), ('quit_month', '=', self.report_month),('escrow_no','=',self.escrow_no.id)])
            mycount = self.env['cloudrent.quit_lease'].search_count([('quit_year', '=', self.report_year), ('quit_month', '=', self.report_month),('escrow_no','=',self.escrow_no.id)])

            mypage = 1
            mypagemod = divmod(mycount, 5)
            if mypagemod[1] > 0:
                mypage = mypagemod[0] + 1
            elif mypagemod[0] > 0 and mypagemod[1] == 0:
                mypage = mypagemod[0]

            if mypage > 1:
                myitem = 6
                for i in range(mypage - 1):
                    target = wb.copy_worksheet(ws)
                    target.title = "第" + str(i + 2) + "頁"
                    myseq = 5
                    for j in range(4):
                        target.cell(column=1, row=myseq).value = myitem
                        myseq = myseq + 1
                        myitem = myitem + 1


            ws.cell(column=1, row=2).value = "業者編號: %s" % self.env.user.escrow_no.bus_no
            ws.cell(column=9, row=2).value = reportdate
            nseq = 5
            mycount = 0
            myworksheet=0
            for rec in myrec:
                A = divmod(mycount, 5)  # 每 sheet 有5筆 紀錄
                if A[0] > 0 and A[1] == 1:  # 換頁
                    myworksheet = myworksheet + 1
                    nseq = 5
                ws = wb.worksheets[myworksheet]
                ws.cell(column=2, row=nseq).value = rec.match_no.match_no     # 媒合編號
                ws.cell(column=2, row=nseq).border = border
                if rec.match_no.case_type=='1':   # 包租包管
                    mycasetype = '■包租包管\n□代租代管'
                elif rec.match_no.case_type=='2':
                    mycasetype = '□包租包管\n■代租代管'
                else:
                    mycasetype = '□包租包管\n□代租代管'
                ws.cell(column=3, row=nseq).value = mycasetype  # 案件類型
                ws.cell(column=3, row=nseq).border = border

                ws.cell(column=4, row=nseq).value = rec.match_no.build_lessee.escrow_man  # 承租人姓名
                ws.cell(column=4, row=nseq).border = border

                ws.cell(column=5, row=nseq).value = rec.match_no.build_lessor.escrow_man  # 出租人姓名
                ws.cell(column=5, row=nseq).border = border

                if rec.match_no.match_start_date:
                    self.env.cr.execute("""select get_twymd('%s')""" % rec.match_no.match_start_date)
                    ws.cell(column=6, row=nseq).value = self.env.cr.fetchone()[0]       # 租約起日
                ws.cell(column=6, row=nseq).border = border

                if rec.match_no.match_end_date:
                    self.env.cr.execute("""select get_twymd('%s')""" % rec.match_no.match_end_date)
                    ws.cell(column=7, row=nseq).value = self.env.cr.fetchone()[0]       # 租約迄日
                ws.cell(column=7, row=nseq).border = border

                if rec.quit_date:
                    self.env.cr.execute("""select get_twymd('%s')""" % rec.quit_date)
                    ws.cell(column=8, row=nseq).value = self.env.cr.fetchone()[0]       # 租約終止日
                ws.cell(column=8, row=nseq).border = border

                if rec.quit_type:
                    ws.cell(column=9, row=nseq).value = rec.quit_type                   # 租約終止項目
                ws.cell(column=9, row=nseq).border = border

                self.env.cr.execute("""select get_applyfor_repair_fee(%d)""" % rec.match_no.id)
                ws.cell(column=10, row=nseq).value = self.env.cr.fetchone()[0]          # 申請修繕費
                ws.cell(column=10, row=nseq).border = border

                self.env.cr.execute("""select get_applyfor_grant_fee(%d)""" % rec.match_no.id)
                ws.cell(column=11, row=nseq).value = self.env.cr.fetchone()[0]          # 申請租金補助
                ws.cell(column=11, row=nseq).border = border

                if rec.substitue_rent:
                    ws.cell(column=12, row=nseq).value = '√'                            # 申請代墊租金
                else:
                    ws.cell(column=12, row=nseq).value = ' '


                nseq = nseq + 1
                mycount = mycount + 1

            wb.save(xlsxnewname)
            myxlsfilename = "[%s]%s-%s房屋租賃契約終止清冊.xlsx" % (self.env.user.escrow_no.guild_no, self.report_year, self.report_month)
            myrec1 = self.env['cloudrent.month_download']
            xlsxfile = open(xlsxnewname, "rb")
            myout = xlsxfile.read()
            myrec1.create({'match_year': self.report_year, 'escrow_no': self.env.user.escrow_no.id,
                          'xls_file': base64.b64encode(myout),'match_month':self.report_month , 'xls_file_name': myxlsfilename})
            os.remove(xlsxnewname)
            myviewid = self.env.ref('cloudrent_escrow.view_cloudrent_month_download_tree')

            return {
                'view_name': 'CloudRentMatchMonthDownload',
                'name': ('房屋租賃契約終止清冊匯出'),
                'type': 'ir.actions.act_window',
                'res_model': 'cloudrent.month_download',
                'view_id': myviewid.id,
                'flags': {'action_buttons': True},
                'view_type': 'form',
                'view_mode': 'tree',
                'target': 'main'}


